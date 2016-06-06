#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import argparse
import numpy as np
from prettytable import PrettyTable
import time
import urllib2
import os.path
import sys
import datetime

__author__ = "Even Langfeldt Friberg"
__copyright__ = "Copyright 2015, Lonely Planet Report Order Helper"
__license__ = "GPL"
__version__ = "1.0"
__maintainer__ = "Even Langfeldt Friberg"
__email__ = "even@evenezer.me"
__status__ = "Production"

def download_lists():
    """
    Downloads current LP complete stock list and incomplete LP outdated
    stocks lists from authors' website. They are used to compare against
    stock found in imported xlsx report.
    """
    try:
        url = "http://folk.uio.no/evenlf/lp-lists/complete_list.txt"
        file_name = url.split('/')[-1]
        u = urllib2.urlopen(url)
        f = open(file_name, 'wb')
        meta = u.info()
        file_size = int(meta.getheaders("Content-Length")[0])
        print "Laster ned: %s byte: %s" % (file_name, file_size)

        file_size_dl = 0
        block_sz = 8192
        while True:
            buffer = u.read(block_sz)
            if not buffer:
                break

            file_size_dl += len(buffer)
            f.write(buffer)
            status = r"%10d  [%3.2f%%]" % (file_size_dl,
                                           file_size_dl * 100. / file_size)
            status = status + chr(8) * (len(status) + 1)
            print status,
        f.close()
    except urllib2.HTTPError, err:
        if err.code == 404:
            print 'FEIL: Filen %s ble ikke funnet. Error 404: Listen kan ha blitt flyttet, eller siden er nede.' % url
            if os.path.isfile('complete_list.txt'):
                try:
                    mtime = os.path.getmtime('complete_list.txt')
                except OSError:
                    mtime = 0
                last_modified_date = datetime.datetime.fromtimestamp(mtime)
                print 'VELG: Hvis du vil fortsette med gammel liste, trykk g etterfulgt av ENTER på tastaturet -- alt annet avslutter.'
                print 'INFO: Gammel liste ble sist endret %s' % last_modified_date
                while True:
                    choice = raw_input("> ")

                    if choice == 'g':
                        print "Bruker gammel liste."
                        break
                    else:
                        print 'Avslutter.'
                        sys.exit(1)
            else:
                print 'INFO: Filen complete-list.txt eksisterer ikke i programmets mappe. Avslutter.'
                sys.exit(1)
        else:
            raise
    except urllib2.URLError, err:
        print 'ERROR: Du er ikke koblet til Internett. Programmet avsluttes.'
        sys.exit(1)

    try:
        url = "http://folk.uio.no/evenlf/lp-lists/old_list.txt"
        file_name = url.split('/')[-1]
        u = urllib2.urlopen(url)
        f = open(file_name, 'wb')
        meta = u.info()
        file_size = int(meta.getheaders("Content-Length")[0])
        print "Laster ned: %s byte: %s\n" % (file_name, file_size)

        file_size_dl = 0
        block_sz = 8192
        while True:
            buffer = u.read(block_sz)
            if not buffer:
                break

            file_size_dl += len(buffer)
            f.write(buffer)
            status = r"%10d  [%3.2f%%]" % (file_size_dl,
                                           file_size_dl * 100. / file_size)
            status = status + chr(8) * (len(status) + 1)
            print status,
            f.close()
    except urllib2.HTTPError, err:
        if err.code == 404:
            print 'FEIL: Filen %s ble ikke funnet. Error 404: Listen kan ha blitt flyttet, eller siden er nede.' % url
            if os.path.isfile('old_list.txt'):
                try:
                    mtime = os.path.getmtime('old_list.txt')
                except OSError:
                    mtime = 0
                last_modified_date = datetime.datetime.fromtimestamp(mtime)
                print 'VELG: Hvis du vil fortsette med gammel liste, trykk g etterfulgt av ENTER på tastaturet -- alt annet avslutter.'
                print 'INFO: Gammel liste ble sist endret %s' % last_modified_date
                while True:
                    choice = raw_input("> ")

                    if choice == 'g':
                        print "Bruker gammel liste."
                        break
                    else:
                        print 'Avslutter.'
                        sys.exit(1)
            else:
                print 'INFO: Filen old_list.txt eksisterer ikke i programmets mappe. Avslutter.'
                sys.exit(1)
        else:
            raise
    except urllib2.URLError, err:
        print 'ERROR: Du er ikke koblet til Internett.'


def read_complete_list(filename, filename2):
    """
    Reads the (downloaded) same-folder stock lists.

    If ISBN is found in both complete_list and old_list, it is not accepted
    in complete_list.
    """
    complete_list = []
    old_list = []
    prohibit_isbn = []

    with open(filename2, 'r') as f:
        for line in f:
            data = []
            words = line.split(";")
            add_isbn = ''.join(words[0].split())
            data.append(add_isbn)
            prohibit_isbn.append(add_isbn)
            data.append(words[1][1:])
            data.append(words[2][1:])
            data.append(words[3][1:-1])
            old_list.append(data)
    f.closed
    print 'INFO: Liste over utdaterte LP-titler lest inn.'

    with open(filename, 'r') as f:
        for line in f:
            data = []
            words = line.split(";")
            add_isbn = ''.join(words[0].split())
            if add_isbn not in prohibit_isbn:
                data.append(add_isbn)
                data.append(words[1][1:])
                data.append(words[2][1:])
                data.append(words[3][1:-1])
                complete_list.append(data)

    f.closed
    print 'INFO: Aktuell LP-katalog lest inn.'

    return complete_list, old_list


def show_results(detected_old_books, detected_good_books, unknown_from_report):
    """
    Prints to stdout outdated titles, current titles and unknown titles found in report.
    """

    print '\nTABELL 1\nFant %s utdaterte (erstattet av ny) LP-titler i din rapport:' % len(
        detected_old_books)
    print '(Dette programmet baserer seg på old_list.txt; du bør også sjekke siste Tabell 3 der\ntitler i rapport uten treff i lister presenteres.)'
    t = PrettyTable(['ISBN', 'Navn', 'Innbinding', 'År', 'Salg totalt',
                     'Beholdning'])
    for b in detected_old_books:
        t.add_row([b[2], b[3], b[4], b[5], b[8], b[10]])
    print t

    print '\nTABELL 2\nFant %s aktuelle bøker fra LP-katalogen nevnt i din rapport:' % len(
        detected_good_books)
    print '(Disse er sortert på beholdning (økende) slik at de øverste normalt er viktigst å bestille.)'
    t_good = PrettyTable(['ISBN', 'Navn', 'Innbinding', 'År', 'Salg totalt',
                          'Beholdning'])
    for b in detected_good_books:
        t_good.add_row([b[2], b[3], b[4], b[5], b[8], b[10]])
    t_good.sortby = 'Beholdning'
    print t_good

    print '\nTABELL 3\nFant %s bøker i din rapport som programmet ikke finner i innlastede godkjent/forbudt-lister:' % len(
        unknown_from_report)
    print '(Disse er typisk ikke-LP-titler og utdaterte LP-titler som gikk OUP før 2015-09.)'
    t_unknown = PrettyTable(['ISBN', 'Navn', 'Innbinding', 'År', 'Salg totalt',
                             'Beholdning'])
    for unknown in unknown_from_report:
        t_unknown.add_row([unknown[2], unknown[3], unknown[4], unknown[5],
                           unknown[8], unknown[10]])
    t_unknown.sortby = 'Beholdning'
    print t_unknown

    qa_number = len(detected_old_books) + len(detected_good_books) + len(
        unknown_from_report)
    return qa_number


def show_not_found(A, np_complete_list):
    """
    Prints to stdout titles from current list and not-yet-published titles not found in report.
    """

    year, month = time.localtime()[0:2]
    if len(str(month)) != 2:
        #print 'fuck'
        #print str(month)
        month = '0' + str(month)
        #print str(month)
    year_month = str(year) + '-' + str(month)
    #print year_month
    #exit(1)
    cnt_missing = 0
    cnt_notpublishedyet = 0
    t_missing = PrettyTable(['ISBN', 'Navn', 'Publikasjonsdato', 'Utgave'])
    t_notpublishedyet = PrettyTable(['ISBN', 'Navn', 'Publikasjonsdato',
                                     'Utgave'])
    A_isbns = []
    for book in A:
        book[2] = int(book[2])
        A_isbns.append(book[2])
    no_titles_complete = np_complete_list.shape[0]
    for i in xrange(no_titles_complete):
        if int(np_complete_list[i][0]) not in A_isbns:
            if year_month >= str(np_complete_list[i][2]):
                t_missing.add_row(
                    [np_complete_list[i][0], np_complete_list[i][1],
                     np_complete_list[i][2], np_complete_list[i][3]])
                cnt_missing += 1
                #print np_complete_list[i][0]
            else:
                t_notpublishedyet.add_row(
                    [np_complete_list[i][0], np_complete_list[i][1],
                     np_complete_list[i][2], np_complete_list[i][3]])
                cnt_notpublishedyet += 1
    print '\nTABELL 4\nFant %s aktuelle bøker fra LP-katalogen som mangler i din rapport:' % cnt_missing
    print '(Du bør slå opp deres ISBN manuelt for å sjekke deres antall i beholdning, evt. generere en'
    print 'rapport som går lenger tilbake i tid. Du mangler trolig enkelte av disse titlene.)'
    t_missing.sortby = 'Publikasjonsdato'
    print t_missing
    print '\nTABELL 5\nFant %s bøker fra LP-katalogen som ikke er publisert ennå [as of %s], og som mangler i din rapport:' % (
        cnt_notpublishedyet, year_month)
    print '(Du bør slå opp deres ISBN manuelt for å sjekke om du har bestilt disse.)'
    t_notpublishedyet.sortby = 'Publikasjonsdato'
    print t_notpublishedyet


if __name__ == "__main__":

    parser = argparse.ArgumentParser(
        description='Lonely Planet Report Order Helper')

    parser.add_argument(
        'infile',
        type=str,
        help='Infile report to read (must be xlsx)')

    parser.add_argument('-v',
                        '--verbose',
                        action='store_true',
                        help='Turn on verbose mode.')

    args = parser.parse_args()

    old_books = [9781741798227]
    permitted_lp_books = [9781742200347]
    detected_old_books = []
    detected_good_books = []
    not_in_report = []
    unknown_from_report = []

    wb = load_workbook(filename=args.infile, use_iterators=True)
    sheet = wb.active
    row_count = sheet.max_row
    column_count = sheet.max_column

    #test
    ##print "TEST"
    #sheet.calculate_dimensions(force=True)
    #exit(1)

    #download_lists()
    complete_list, old_list = read_complete_list('complete_list.txt',
                                                 'old_list.txt')
    np_complete_list = np.array(complete_list)
    np_old_list = np.array(old_list)

    upperleftcell = 'A3'
    lowerrightcell = 'K' + str(row_count - 2)
    A = np.array([[i.value for i in j] for j in sheet[upperleftcell:
                                                      lowerrightcell]])

    for book in A:
        book[2] = int(book[2])
        if book[2] in np_complete_list[:, 0].astype(int):
            if book[2] in np_old_list[:, 0].astype(int):
                detected_old_books.append(book)
            else:
                # overwrites report-title with list-title, ok
                name_index = np.where(np_complete_list[:, 0] == str(book[2]))
                name_index = name_index[-1][0]
                book[3] = np_complete_list[name_index][1]
                detected_good_books.append(book)
        else:
            if str(book[2]) in np_old_list[:, 0]:
                detected_old_books.append(book)
            else:
                unknown_from_report.append(book)

    qa_number = show_results(detected_old_books, detected_good_books,
                             unknown_from_report)
    show_not_found(A, np_complete_list)

    if int(qa_number) == int(row_count - 4):
        print 'INFO: Alle %s titler i din rapport ble klassifisert og plassert i en tabell.' % str(
            row_count - 4)
    else:
        print 'ERROR: Bare %s titler fra din rapport ble klassifisert og plassert i en tabell,\n men rapporten inneholder %s titler! Kontakt utvikleren på even@evenezer.me og legg med rapport og utskrift!' % (
            qa_number, (row_count - 4))
